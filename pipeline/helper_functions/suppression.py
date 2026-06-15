from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .settings import CONFIG

DISCLOSIVE_THRESHOLD = 5
UNRELIABLE_THRESHOLD = 10
SHADE = PatternFill(start_color='999999', end_color='999999', fill_type="solid")
DISCLOSIVE_SYMBOL = "[c]"
UNRELIABLE_SYMBOL = "[u]"


def run_suppression_pipeline(paths):
    for path in paths:
        print(f"\nSuppressing {path}")
        suppress_data(path)


def suppress_data(path):
    file = load_workbook(path)
    iterate_through_sheets(file)
    file.save(path)


def iterate_through_sheets(file):
    for sheet in file.sheetnames:
        # Set sheet name to active sheet
        active_sheet = file[sheet]

        # Get the max row value
        max_row = active_sheet.max_row

        # Get column names
        headers = {cell.value: cell.column for cell in active_sheet[1]}

        # Find all Raw_YYYY columns
        raw_cols = [h for h in headers if h.startswith("Raw_")]

        raw_col_idxs = iterate_through_yearly_data(active_sheet, headers, max_row, raw_cols)

        if CONFIG['delete_raw_data']:
            delete_raw_data(active_sheet, raw_col_idxs)


def delete_raw_data(active_sheet, raw_col_idxs):
    for raw_col_idx in sorted(raw_col_idxs, reverse=True):
        # Reverse order as indices will shift after deletion otherwise
        active_sheet.delete_cols(raw_col_idx)


def iterate_through_yearly_data(active_sheet, headers, max_row, raw_cols):
    raw_col_idxs = []

    for raw_col_name in raw_cols:
        raw_col_idx = headers[raw_col_name]
        raw_col_idxs.append(raw_col_idx)
        raw_values = [active_sheet.cell(row=row, column=raw_col_idx).value for row in range(2, max_row + 1)]

        year = raw_col_name.split("_")[1]

        weighted_cols = [f"{year}"]
        if int(year) in CONFIG['sampling_files']:
            weighted_cols.extend([f'{year} Confidence Interval lower', f'{year} Confidence Interval upper'])

        for row, raw_value in enumerate(raw_values, start=2):
            raw_cell = active_sheet.cell(row=row, column=raw_col_idx)

            if (CONFIG['replace_disclosive_data'] and isinstance(raw_cell.value, (int, float))
                    and raw_cell.value < DISCLOSIVE_THRESHOLD):
                _suppress_weighted_cells(headers, weighted_cols, row, active_sheet, suppression_type='disclosive')

            elif (CONFIG['shade_unreliable_data'] and isinstance(raw_cell.value, (int, float))
                  and raw_cell.value < UNRELIABLE_THRESHOLD):
                _suppress_weighted_cells(headers, weighted_cols, row, active_sheet, suppression_type='unreliable')

    return raw_col_idxs


def _suppress_weighted_cells(headers, weighted_cols, row, active_sheet, suppression_type):
    for col in weighted_cols:
        weighted_col_idx = headers[col]
        weighted_cell = active_sheet.cell(row=row, column=weighted_col_idx)

        if suppression_type == 'disclosive':
            weighted_cell.value = DISCLOSIVE_SYMBOL
        else:
            weighted_cell.fill = SHADE
